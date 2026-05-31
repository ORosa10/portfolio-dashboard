#!/usr/bin/env python3
"""Generate dashboard JavaScript data files from an Excel workbook.

Usage:
  python scripts/generate_dashboard_data.py --excel path/to/workbook.xlsx --mapping config/excel-mapping.json
"""

from __future__ import annotations

import argparse
import json
import math
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


@dataclass
class Period:
    key: str
    month: str
    year: int
    column: int


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True, help="Path to the source Excel workbook")
    parser.add_argument("--mapping", default="config/excel-mapping.json", help="Path to mapping JSON")
    parser.add_argument("--out-dir", default=".", help="Repository root / output base directory")
    return parser.parse_args()


def normalise_number(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) if isinstance(value, float) else False:
            return None
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(" ", "").replace(",", ".")
        if cleaned in {"", "-", "n/a", "N/A"}:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def js_number(value: Optional[float]) -> Any:
    if value is None:
        return None
    rounded = round(float(value), 4)
    if rounded.is_integer():
        return int(rounded)
    return rounded


def infer_periods(ws: Any, date_row: int, start_col: int, fallback_start_period: str) -> List[Period]:
    periods: List[Period] = []
    fallback_year, fallback_month = [int(part) for part in fallback_start_period.split("-")]

    for col in range(start_col, ws.max_column + 1):
        cell_value = ws.cell(row=date_row, column=col).value
        period_key: Optional[str] = None

        if isinstance(cell_value, datetime):
            period_key = f"{cell_value.year}-{cell_value.month:02d}"
        elif isinstance(cell_value, str):
            raw = cell_value.strip()
            for fmt in ("%b-%y", "%b %y", "%b-%Y", "%b %Y", "%Y-%m"):
                try:
                    dt = datetime.strptime(raw, fmt)
                    period_key = f"{dt.year}-{dt.month:02d}"
                    break
                except ValueError:
                    pass

        if period_key is None:
            month_index = fallback_month + (col - start_col) - 1
            year = fallback_year + month_index // 12
            month = month_index % 12 + 1
            period_key = f"{year}-{month:02d}"

        year = int(period_key[:4])
        month_number = int(period_key[5:7])
        periods.append(Period(key=period_key, month=MONTHS[month_number - 1], year=year, column=col))

    return periods


def read_metric_row(ws: Any, row_number: int, periods: List[Period]) -> Dict[str, Optional[float]]:
    return {period.key: normalise_number(ws.cell(row=row_number, column=period.column).value) for period in periods}


def build_records(periods: List[Period], actuals: Dict[str, Optional[float]], budgets: Dict[str, Optional[float]]) -> List[Dict[str, Any]]:
    records = []
    for period in periods:
      ly_key = f"{period.year - 1}-{period.key[5:7]}"
      records.append({
          "period": period.key,
          "month": period.month,
          "year": period.year,
          "actual": js_number(actuals.get(period.key)),
          "budget": js_number(budgets.get(period.key)),
          "ly": js_number(actuals.get(ly_key)),
      })
    return records


def subtract_series(left: Dict[str, Optional[float]], right: Dict[str, Optional[float]]) -> Dict[str, Optional[float]]:
    out: Dict[str, Optional[float]] = {}
    for key, left_value in left.items():
        right_value = right.get(key)
        out[key] = None if left_value is None or right_value is None else left_value - right_value
    return out


def build_metric_series(metric_id: str, metric_config: Dict[str, Any], all_actuals: Dict[str, Dict[str, Optional[float]]], all_budgets: Dict[str, Dict[str, Optional[float]]], periods: List[Period]) -> List[Dict[str, Any]]:
    if metric_config.get("formula") == "adjusted_ebitda - capex":
        actuals = subtract_series(all_actuals["adjusted_ebitda"], all_actuals["capex"])
        budgets = subtract_series(all_budgets["adjusted_ebitda"], all_budgets["capex"])
    else:
        actuals = all_actuals[metric_id]
        budgets = all_budgets[metric_id]
    return build_records(periods, actuals, budgets)


def write_js(path: Path, assignment: str, payload: Dict[str, Any], header: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = header + "\n" + f"window.{assignment} = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n"
    path.write_text(text, encoding="utf-8")


def main() -> None:
    args = parse_args()
    repo_root = Path(args.out_dir)
    mapping_path = Path(args.mapping)
    mapping = json.loads(mapping_path.read_text(encoding="utf-8"))

    workbook_path = Path(args.excel)
    wb = load_workbook(workbook_path, data_only=True)

    actual_sheet = mapping["workbook"]["actual_sheet"]
    budget_sheet = mapping["workbook"]["budget_sheet"]
    date_row = int(mapping["workbook"]["date_row"])
    start_col = column_index_from_string(mapping["workbook"]["start_column"])
    start_period = mapping["workbook"].get("start_period", "2020-01")

    ws_actual = wb[actual_sheet]
    ws_budget = wb[budget_sheet]
    periods = infer_periods(ws_actual, date_row, start_col, start_period)

    all_actuals: Dict[str, Dict[str, Optional[float]]] = {}
    all_budgets: Dict[str, Dict[str, Optional[float]]] = {}

    for metric_id, metric_config in mapping["metrics"].items():
        if "formula" in metric_config:
            continue
        all_actuals[metric_id] = read_metric_row(ws_actual, int(metric_config["actual_row"]), periods)
        all_budgets[metric_id] = read_metric_row(ws_budget, int(metric_config["budget_row"]), periods)

    metrics = mapping["metrics"]

    revenues_payload = {
        "metricId": "revenues",
        "label": metrics["revenues"]["label"],
        "sourceLabel": metrics["revenues"]["source_label"],
        "unit": metrics["revenues"]["unit"],
        "actualSource": {"sheet": actual_sheet, "dateRow": date_row, "valueRow": metrics["revenues"]["actual_row"]},
        "budgetSource": {"sheet": budget_sheet, "dateRow": date_row, "valueRow": metrics["revenues"]["budget_row"]},
        "lyRule": "actual_minus_12m",
        "records": build_metric_series("revenues", metrics["revenues"], all_actuals, all_budgets, periods),
    }

    operating_ids = [metric_id for metric_id, cfg in metrics.items() if cfg["category"] == "operating"]
    operating_payload = {
        "metadata": {
            "unit": mapping["workbook"].get("unit", "CZKm"),
            "sourceWorkbook": workbook_path.name,
            "actualSheet": actual_sheet,
            "budgetSheet": budget_sheet,
            "dateRow": date_row,
            "startColumn": mapping["workbook"]["start_column"],
            "lyRule": "actual_minus_12m",
        },
        "metrics": {
            metric_id: {
                "label": metrics[metric_id]["label"],
                "sourceLabel": metrics[metric_id]["source_label"],
                "actualRow": metrics[metric_id].get("actual_row"),
                "budgetRow": metrics[metric_id].get("budget_row"),
                "formula": metrics[metric_id].get("formula"),
                "unit": metrics[metric_id]["unit"],
                "lowerIsBetter": metrics[metric_id]["lower_is_better"],
            }
            for metric_id in operating_ids
        },
        "records": {
            metric_id: build_metric_series(metric_id, metrics[metric_id], all_actuals, all_budgets, periods)
            for metric_id in operating_ids
        },
    }

    balance_ids = [metric_id for metric_id, cfg in metrics.items() if cfg["category"] == "balance_sheet"]
    balance_payload = {
        "metadata": {
            "unit": mapping["workbook"].get("unit", "CZKm"),
            "sourceWorkbook": workbook_path.name,
            "actualSheet": actual_sheet,
            "budgetSheet": budget_sheet,
            "dateRow": date_row,
            "startColumn": mapping["workbook"]["start_column"],
            "lyRule": "actual_minus_12m",
        },
        "metrics": {
            metric_id: {
                "label": metrics[metric_id]["label"],
                "sourceLabel": metrics[metric_id]["source_label"],
                "actualRow": metrics[metric_id].get("actual_row"),
                "budgetRow": metrics[metric_id].get("budget_row"),
                "unit": metrics[metric_id]["unit"],
                "lowerIsBetter": metrics[metric_id]["lower_is_better"],
            }
            for metric_id in balance_ids
        },
        "records": {
            metric_id: build_metric_series(metric_id, metrics[metric_id], all_actuals, all_budgets, periods)
            for metric_id in balance_ids
        },
    }

    outputs = mapping["outputs"]
    write_js(repo_root / outputs["revenues"], "RevenueData", revenues_payload, "/* Auto-generated from Excel. Do not edit manually. */")
    write_js(repo_root / outputs["operating_metrics"], "OperatingMetricData", operating_payload, "/* Auto-generated from Excel. Do not edit manually. */")
    write_js(repo_root / outputs["balance_sheet"], "BalanceSheetData", balance_payload, "/* Auto-generated from Excel. Do not edit manually. */")

    print("Generated dashboard data files:")
    for output in outputs.values():
        print(f"- {output}")


if __name__ == "__main__":
    main()
